import os
import re
import time
import jieba
import socket
import ipaddress
import requests
import tldextract
from collections import Counter
from openpyxl import Workbook
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from requests.exceptions import RequestException
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


ip_pattern = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
ip_port_pattern = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}:\d+$")
domain_port_pattern = re.compile(r"^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}:\d+$")


def download_china_mainland_data(file_path, ip_data_url):
    """下载中国大陆境内IP数据"""
    response = requests.get(ip_data_url, stream=True, timeout=(10, 3600))
    with open(file_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)


def load_china_mainland_data(file_path, ip_data_url):
    """优先从缓存文件中加载数据"""
    if not os.path.exists(file_path):
        download_china_mainland_data(file_path, ip_data_url)
    with open(file_path, 'r') as file:
        return file.read()


def parse_china_mainland_data(ip_data):
    """解析IP段数据"""
    cn_ip_ranges = []
    for line in ip_data.splitlines():
        cn_ip_ranges.append(ipaddress.ip_network(line))
    return cn_ip_ranges


def is_china_mainland_ip(ip, cn_ip_ranges):
    """判断IP是否属于中国大陆境内"""
    if not cn_ip_ranges:
        return '未知'
    ip_address = ipaddress.ip_address(ip)
    for network in cn_ip_ranges:
        if ip_address in network:
            return '是'
    return '否'


def resolve_domain(domain, max_retries=3, timeout=3):
    """解析域名的IP地址，如果输入的域名是IP、IP:端口或域名:端口形式的，则正确解析IP"""

    # 检查是否为纯 IP 地址
    if ip_pattern.match(domain):
        # 直接返回 IP 地址
        return domain

    # 检查是否为 IP:端口 的格式
    if ip_port_pattern.match(domain):
        # 提取并返回 IP 部分
        return domain.split(':')[0]

    # 检查是否为 域名:端口 的格式
    if domain_port_pattern.match(domain):
        # 提取并保留域名部分（去掉端口）
        domain = domain.split(':')[0]

    # 否则，进行域名解析
    retries = 0
    while retries < max_retries:
        try:
            # 设置超时时间
            socket.setdefaulttimeout(timeout)
            ip_address = socket.gethostbyname(domain)
            return ip_address
        except (socket.timeout, socket.gaierror) as e:
            retries += 1
            time.sleep(1)  # 等待一段时间后重试
    return ''


def extract_keywords(text, max_num=20):
    """切词并选取出现最多的词语"""
    # 使用jieba分词
    words = jieba.cut(text)

    # 统计词频
    word_count = Counter(words)

    # 过滤掉单字和标点符号
    filtered_words = [(word, count) for word, count in word_count.items() if len(word) > 1]

    # 获取前top_n的关键词
    keywords = sorted(filtered_words, key=lambda x: x[1], reverse=True)[:max_num]

    return [keyword[0] for keyword in keywords]


def calculate_match_and_ratio(input_words, keyword_combinations, all_keywords, max_num=10):
    """计算匹配的关键词组合和出现的词语在所有关键词中的占比"""
    if not input_words:
        return [], 0.0

    matched_combinations = []
    input_word_set = set(input_words)
    matched_words = set()

    for combination in keyword_combinations:
        if all(word in input_word_set for word in combination):
            matched_combinations = combination
            break

    for word in input_words[:max_num]:
        if word in all_keywords:
            matched_words.add(word)

    ratio = round(len(matched_words) / len(input_words[:max_num]), 2) if input_words else 0.0

    return matched_combinations, ratio


def load_keywords(file_path):
    """加载关键词"""
    with open(file_path, 'r', encoding='utf-8') as f:
        keyword_combinations = []
        all_keywords = set()
        for line in f:
            keywords = line.strip().split('\t')
            keyword_combinations.append(keywords)
            all_keywords.update(keywords)
        return keyword_combinations, all_keywords


def filter_long_lines(content, max_length=1000):
    """过滤过长的行避免影响正则匹配的效率"""
    lines = content.splitlines()
    filtered_lines = [line for line in lines if len(line) <= max_length]
    return "\n".join(filtered_lines)


def save_data_to_excel(file_path, headers, data):
    """将数据写入Excel文件"""
    wb_save = Workbook()
    ws_save = wb_save.active
    # 写入标题
    for col_index, header in enumerate(headers, start=1):
        ws_save.cell(row=1, column=col_index, value=header)

    # 写入数据
    for row_index, row_data in enumerate(data, start=2):
        for col_index, cell_value in enumerate(row_data, start=1):
            ws_save.cell(row=row_index, column=col_index, value=cell_value)

    wb_save.save(file_path)


def calculate_similarity(list1, list2):
    """计算两个关键词列表的相似度"""
    set1, set2 = set(list1), set(list2)
    intersection = len(set1 & set2)
    union = min(len(set1), len(set2))
    similarity = round(intersection / union if union else 0, 2)
    return similarity


def extract_js_urls(html_content, domain):
    """从HTML中提取所有JS文件的URL，并处理不完整的URL和省略协议的URL"""

    # 解析HTML内容
    soup = BeautifulSoup(html_content, 'lxml')
    js_urls = []

    # 提取所有的<script>标签
    scripts = soup.find_all('script', src=True)

    for script in scripts:
        js_url = script['src']

        # 如果URL以 // 开头，补全为 http://
        if js_url.startswith('//'):
            js_url = 'http:' + js_url
        # 如果URL是相对路径，拼接域名
        elif not js_url.startswith(('http://', 'https://')):
            js_url = urljoin(f'http://{domain}', js_url)

        js_urls.append(js_url)

    return js_urls


def fetch_js_content(url, max_retries=3, timeout=5):
    """请求JS文件的URL以获取其内容"""
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url, timeout=timeout)

            if response.status_code == 200:
                return response.text

        except RequestException:
            pass

        retries += 1
        time.sleep(1)

    return None


def get_main_domain(url_or_domain):
    """提取URL或域名的主域名"""
    # 使用tldextract来分解域名或URL
    extracted = tldextract.extract(url_or_domain)
    # 组合主域名和顶级域名
    return f"{extracted.domain}.{extracted.suffix}"


def compare_main_domains(url_or_domain1, url_or_domain2):
    """比较两个URL或域名的主域名是否一致"""
    main_domain1 = get_main_domain(url_or_domain1)
    main_domain2 = get_main_domain(url_or_domain2)
    return main_domain1 == main_domain2


def extract_text_from_html(html_content, text_separator=' '):
    """从HTML内容中提取所有文本，包括title、keywords和description"""
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取title
    title_tags = soup.find_all('title')
    title = text_separator.join(
        [title_tag.get_text(separator=text_separator, strip=True) for title_tag in title_tags]).strip()

    # 提取keywords
    keywords_tag = soup.find('meta', attrs={'name': 'keywords'})
    keywords = keywords_tag['content'].strip() if keywords_tag and 'content' in keywords_tag.attrs else ''

    # 提取description
    description_tag = soup.find('meta', attrs={'name': 'description'})
    description = description_tag['content'].strip() if description_tag and 'content' in description_tag.attrs else ''

    # 提取页面正文文本
    body_text = soup.get_text(separator=text_separator, strip=True)

    # 查找所有设置为hidden的元素
    hidden_elements = soup.find_all(
        lambda tag: tag.has_attr('hidden')
        or ('style' in tag.attrs and 'display:none' in tag['style'].replace(' ', '')))

    # 提取hidden元素中的文本
    hidden_texts = text_separator.join(
        [element.get_text(separator=text_separator, strip=True) for element in hidden_elements])

    # 合并所有文本内容
    combined_text = text_separator.join(filter(None, [title, keywords, description, body_text, hidden_texts]))

    return combined_text


def compute_similarity(text1, text2):
    """计算两段文本内容的相似程度"""

    if not text1 or not text2:
        return 0

    # 使用 jieba 分词
    text1_tokenized = ' '.join(jieba.lcut(text1))
    text2_tokenized = ' '.join(jieba.lcut(text2))

    if not text1_tokenized.strip() or not text2_tokenized.strip():
        return 0

    try:
        # 使用TF-IDF向量化文本
        vectorizer = TfidfVectorizer(stop_words=None)
        tfidf_matrix = vectorizer.fit_transform([text1_tokenized, text2_tokenized])

        # 计算Cosine相似度
        similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])
    except Exception:
        return 0

    return round(similarity[0][0], 2)


def partial_match(pattern, text, max_length=20000, slice_length=10000):
    """
    如果字符串长度超过一定长度，则针对字符串的开头和结尾进行正则表达式匹配，避免直接对超长文本进行匹配，导致耗时过长。

    参数:
    - pattern: 正则表达式的模式。
    - text: 需要匹配的字符串。
    - max_length: 如果字符串长度超过该值，则只对开头和结尾进行匹配。默认值为20000。
    - slice_length: 匹配开头和结尾部分的子字符串长度。默认值为10000。

    返回:
    - search_result: 匹配结果。
    """
    if len(text) > max_length:
        # 提取字符串的开头和结尾部分
        start_text = text[:slice_length]
        end_text = text[-slice_length:]
        # 对开头和结尾进行匹配
        search_result = pattern.search(start_text)
        if not search_result:
            search_result = pattern.search(end_text)
    else:
        # 如果长度不超过 max_length，对整个字符串进行匹配
        search_result = re.findall(pattern, text)

    return search_result


def remove_control_characters(value):
    """移除字符串中的控制字符"""
    if isinstance(value, str):
        # 仅保留常见的字符（Unicode 码点大于等于32）
        return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)
    return value

